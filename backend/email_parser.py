import re
import logging
import io
from typing import Dict, List, Optional, Tuple, BinaryIO
from email import policy
from email.parser import BytesParser
from email.header import decode_header
import html2text
import tempfile
import os
from typing import Dict, Any
from collections import Counter


# Служебные проверки установленных библиотек для корректной работы парсера
try:
    import extract_msg
    MSG_SUPPORT = True
except ImportError:
    MSG_SUPPORT = False
    logging.warning("extract-msg not installed. .msg files will not be supported.")

try:
    import PyPDF2
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
    logging.warning("PyPDF2 not installed. PDF files will not be supported.")

try:
    import docx
    DOCX_SUPPORT = True
except ImportError:
    DOCX_SUPPORT = False
    logging.warning("python-docx not installed. DOCX files will not be supported.")

try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    logging.warning("openpyxl not installed. Excel files will not be supported.")

logger = logging.getLogger(__name__)


class EmailParser:
    """
    Класс парсер писем для извлечения текстового содержимого .eml и .msg файлов
    """

    def __init__(self):
        self.html_converter = html2text.HTML2Text()
        # Настройки для html2text парсера
        self.html_converter.ignore_links = False # Игнорирование ссылок
        self.html_converter.ignore_images = True # Игнорирование изображений
        self.html_converter.ignore_tables = True  # Игнорирование таблиц
        self.html_converter.ignore_emphasis = True  # Игнорирование *курсивf* и **жирного**
        self.html_converter.body_width = True # Количество пустых строк между абзацами (1)
        self.html_converter.single_line_break = True  # Запрет на перенос строк
        self.html_converter.mark_code = False  #  Запрет на обрамление кода в ```

    def decode_email_header(self, header: Optional[str]) -> str:
        """
        Функция декодирования заголовков email с учетом различных кодировок
        """
        if not header:
            return ""

        try:
            decoded_parts = decode_header(header)
            result_parts = []

            for content, encoding in decoded_parts:
                if isinstance(content, bytes):
                    if encoding:
                        try:
                            result_parts.append(content.decode(encoding))
                        except (UnicodeDecodeError, LookupError):
                            result_parts.append(content.decode('utf-8', errors='ignore'))
                    else:
                        result_parts.append(content.decode('utf-8', errors='ignore'))
                else:
                    result_parts.append(str(content))

            return ''.join(result_parts)
        except Exception as e:
            logger.error(f"Ошибка при декодировании заголовка письма: {e}")
            return str(header) if header else ""

    def parse_eml(self, file: bytes, filename: str = "unknown.eml") -> Dict:
        """Парсит .eml файл и возвращает структурированную информацию"""
        try:
            # Создаем байтовый поток
            byte_stream = io.BytesIO(file)
            msg = BytesParser(policy=policy.default).parse(byte_stream)

            # Извлекаем базовые метаданные
            result = {
                'subject': self.decode_email_header(msg.get('subject')),
                'from': self.decode_email_header(msg.get('from')),
                'to': self.decode_email_header(msg.get('to')),
                'cc': self.decode_email_header(msg.get('cc')),
                'date': self.decode_email_header(msg.get('date')),
                'body_plain': '',
                'body_html': '',
                'attachments': [],
                'headers': {},
                'filename': filename
            }

            # Извлекаем дополнительные заголовки
            for header in ['Message-ID', 'References', 'In-Reply-To', 'X-Mailer']:
                if header in msg:
                    result['headers'][header] = self.decode_email_header(msg[header])

            # Рекурсивно обходим части письма
            self._process_email_parts(msg, result)

            # Если нет текстового тела, пытаемся извлечь из html
            if not result['body_plain'] and result['body_html']:
                result['body_plain'] = self.html_converter.handle(result['body_html'])

            logger.info(f"Успешно распарсен .eml файл: {result['subject'][:50]}...")
            return result

        except Exception as e:
            logger.error(f"Ошибка во время парсинга .eml файла: {e}")
            raise ValueError(f"Ошибка во время парсинга .eml файла: {e}")

    def clean_text(self, text: str) -> str:
        """
        Финальная очистка текста от ненужных данных:
        - Unicode zero-width spaces
        - HTML/CSS остатки
        - Служебные символы
        """
        if not text:
            return ""
        
        invisible_chars = [
            '\u200C',
            '\u200B',
            '\u200D',
            '\uFEFF',
            '\u2060',
            
            # Пустые символы Брайля (в твоем примере)
            '\u2800',
            
            # Другие необычные пробелы
            '\u00A0',
            '\u202F',
            '\u205F',
            '\u3000',
            
            # Разные типы пробелов
            '\u2000',
            '\u2001',
            '\u2002',
            '\u2003',
            '\u2004',
            '\u2005',
            '\u2006',
            '\u2007',
            '\u2008',
            '\u2009',
            '\u200A',
            
            # Разделители строк/параграфов
            '\u2028',
            '\u2029',
        ]
        
        # Удаляем все Unicode символы
        for char in invisible_chars:
            text = text.replace(char, ' ')

        # Удаляем повторяющиеся символы (например, много запятых)
        text = re.sub(r'[,;]{3,}', ' ', text)

        # Пустые скобки
        text = re.sub(r'\[\s*\]', ' ', text)
        text = re.sub(r'\(\s*\)', ' ', text)
        text = re.sub(r'\{\s*\}', ' ', text)
        text = re.sub(r'[\[\]\(\){}]', ' ', text)

        # Нормализуем пробелы и переносы
        text = re.sub(r'\r\n?', '\n', text)  
        text = re.sub(r'[ \t]+', ' ', text)  
        text = re.sub(r'\n{3,}', '\n\n', text)
        
        # Убираем строки, которые состоят в основном из спецсимволов
        lines = []
        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue
                
            # Считаем буквы vs не-буквы
            letters = len(re.findall(r'[a-zA-Zа-яА-Я]', line))
            total = len(line)
            
            # Если в строке меньше 30% букв или она очень короткая - пропускаем
            if total > 3 and (letters / total > 0.3 or letters >= 3):
                lines.append(line)
        
        text = '\n'.join(lines)
        
        # Финальная чистка
        text = re.sub(r'\s+', ' ', text)
        
        return text

    def extract_and_remove_urls(self, text: str, placeholder: str = "") -> tuple[str, list[str]]:
        """
        Удаляет все URL из текста и возвращает очищенный текст + список ссылок.
        
        Args:
            text: исходный текст со ссылками
            placeholder: чем заменять ссылки в тексте
            
        Returns:
            Tuple[очищенный_текст, список_ссылок]
        """
        if not text:
            return text, []
        
        # Паттерн для поиска ссылок
        url_patterns = [
            r'(https?://\S+|www\.\S+)',
            r'https?://\S+|www\.\S+',
        ]
        
        urls_found = []
        # Находим все ссылки
        for pattern in url_patterns:
            urls_found.extend(re.findall(pattern, text, flags=re.IGNORECASE))
        
        if not urls_found:
            return text, []
        
        # Заменяем ссылки в тексте
        result = text
        for pattern in url_patterns:
            result = re.sub(pattern, placeholder, result, flags=re.IGNORECASE)
        
        # Нормализуем пробелы (если много ссылок подряд)
        result = re.sub(rf'({re.escape(placeholder)})+', placeholder, result)
        result = re.sub(r'\s+', ' ', result)
        
        return result.strip(), urls_found

    def parse_msg(self, file: bytes, filename: str = "unknown.msg") -> Dict:
        """Парсит .msg файл и возвращает структурированную информацию"""
        if not MSG_SUPPORT:
            raise ImportError("extract-msg library is not installed. Install with: pip install extract-msg")

        try:
            # Сохраняем байты во временный файл для extract_msg
            with tempfile.NamedTemporaryFile(suffix='.msg', delete=False) as tmp_file:
                tmp_file.write(file)
                tmp_file_path = tmp_file.name

            try:
                msg = extract_msg.Message(tmp_file_path)

                result = {
                    'subject': msg.subject or '',
                    'from': msg.sender or '',
                    'to': msg.to or '',
                    'cc': msg.cc or '',
                    'date': str(msg.date) if msg.date else '',
                    'body_plain': '',
                    'body_html': '',
                    'attachments': [],
                    'headers': {},
                    'filename': filename
                }

                # Получаем тело письма
                if msg.body:
                    result['body_plain'] = msg.body

                if msg.htmlBody:
                    result['body_html'] = msg.htmlBody
                    if not result['body_plain']:
                        result['body_plain'] = self.html_converter.handle(msg.htmlBody)

                # Извлекаем вложения
                for attachment in msg.attachments:
                    if hasattr(attachment, 'longFilename') and attachment.longFilename:
                        attachment_filename = attachment.longFilename
                    elif hasattr(attachment, 'filename') and attachment.filename:
                        attachment_filename = attachment.filename
                    else:
                        attachment_filename = f"attachment_{len(result['attachments'])}"

                    # Получаем данные вложения
                    if hasattr(attachment, 'data'):
                        attachment_data = attachment.data
                    else:
                        # Пытаемся получить данные другими способами
                        try:
                            if hasattr(attachment, '_getStream'):
                                attachment_data = attachment._getStream('__substg1.0_37010102')
                            else:
                                attachment_data = b''
                        except:
                            attachment_data = b''

                    # Сохраняем информацию о вложении
                    attachment_info = {
                        'filename': attachment_filename,
                        'data': attachment_data
                    }
                    result['attachments'].append(attachment_info)

                # Очищаем текст
                if result['body_plain']:
                    result['body_plain'] = re.sub(r'\s+', ' ', result['body_plain']).strip()

                logger.info(f"Успешно распарсен .msg файл: {result['subject'][:50]}...")
                return result

            finally:
                # Всегда удаляем временный файл
                try:
                    os.unlink(tmp_file_path)
                except:
                    pass

        except Exception as e:
            logger.error(f"Ошибка во время парсинга .msg файла: {e}")
            raise ValueError(f"Ошибка во время парсинга .msg файла: {e}")

    def _process_email_parts(self, part, result: Dict):
        """
        Рекурсивно обрабатывает части письма (для .eml)
        """
        if part.is_multipart():
            for subpart in part.iter_parts():
                self._process_email_parts(subpart, result)
        else:
            content_type = part.get_content_type()
            content_disposition = str(part.get("Content-Disposition", ""))

            # Проверяем, является ли часть вложением
            if "attachment" in content_disposition or part.get_filename():
                filename = part.get_filename()
                if filename:
                    filename = self.decode_email_header(filename)
                
                attachment_info = {
                    'filename': filename or f"attachment_{len(result['attachments'])}",
                    'content_type': content_type,
                    'data': part.get_payload(decode=True)
                }
                result['attachments'].append(attachment_info)
            else:
                # Часть тела письма
                payload = part.get_payload(decode=True)
                charset = part.get_content_charset() or 'utf-8'

                try:
                    if payload:
                        text = payload.decode(charset, errors='replace')
                        if content_type == 'text/plain':
                            result['body_plain'] += text + "\n"
                        elif content_type == 'text/html':
                            result['body_html'] += text + "\n"
                except (UnicodeDecodeError, LookupError) as e:
                    logger.warning(f"Ошибка декодирования части кода, связанной с кодировкой символов {charset}: {e}")
                    try:
                        text = payload.decode('utf-8', errors='replace')
                        if content_type == 'text/plain':
                            result['body_plain'] += text + "\n"
                        elif content_type == 'text/html':
                            result['body_html'] += text + "\n"
                    except:
                        pass

    def extract_text_from_pdf(self, file: bytes, filename: str = "unknown.pdf") -> str:
        """Извлекает текст из PDF файла"""
        if not PDF_SUPPORT:
            return f"PDF содержимое недоступно - установите PyPDF2: {filename}"
        try:
            text = ""
            pdf_stream = io.BytesIO(file)

            pdf_reader = PyPDF2.PdfReader(pdf_stream)

            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text.strip()
        except Exception as e:
            logger.error(f"Ошибка при извлечении текста из PDF {filename}: {e}")
            raise ValueError(f"Ошибка при извлечении текста из PDF {filename}: {e}")

    def extract_text_from_docx(self, file: bytes, filename: str = "unknown.docx") -> str:
        """Извлекает текст из DOCX файла"""
        if not DOCX_SUPPORT:
            return f"[DOCX содержимое недоступно - установите python-docx: {filename}]"

        try:
            doc_stream = io.BytesIO(file)
            doc = docx.Document(doc_stream)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text.strip()
        except Exception as e:
            logger.error(f"Ошибка при извлечении текста из DOCX {filename}: {e}")
            raise ValueError(f"Ошибка при извлечении текста из DOCX {filename}: {e}")

    def extract_text_from_excel(self, file: bytes, filename: str = "unknown.xlsx") -> str:
        """Извлекает текст из Excel файла"""
        if not EXCEL_SUPPORT:
            return f"[Excel содержимое недоступно - установите openpyxl: {filename}]"

        try:
            excel_stream = io.BytesIO(file)
            wb = load_workbook(excel_stream, read_only=True, data_only=True)
            text_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_text = []

                for row in ws.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        sheet_text.append(" | ".join(row_text))

                if sheet_text:
                    text_parts.append(f"--- Лист: {sheet_name} ---")
                    text_parts.extend(sheet_text)

            return "\n".join(text_parts).strip()
        except Exception as e:
            logger.error(f"Ошибка при извлечении текста из Excel {filename}: {e}")
            raise ValueError(f"Ошибка при извлечении текста из Excel {filename}: {e}")

    def extract_text_from_csv(self, file: bytes, filename: str = "unknown.csv") -> str:
        """Извлекает текст из CSV файла"""
        try:
            import csv
            text_parts = []
            csv_stream = io.BytesIO(file)
            
            # Пытаемся определить кодировку
            sample = csv_stream.read(1024)
            csv_stream.seek(0)
            
            # Пробуем разные кодировки
            for encoding in ['utf-8', 'cp1251', 'koi8-r', 'iso-8859-5']:
                try:
                    csv_stream.seek(0)
                    decoded = csv_stream.read().decode(encoding)
                    csv_text = io.StringIO(decoded)
                    
                    # Пытаемся определить разделитель
                    sniffer = csv.Sniffer()
                    try:
                        dialect = sniffer.sniff(decoded[:1024])
                    except:
                        dialect = csv.excel
                    
                    csv_reader = csv.reader(csv_text, dialect)
                    for row in csv_reader:
                        text_parts.append(" | ".join(row))
                    
                    break  # Успешно декодировали
                except:
                    continue
            
            return "\n".join(text_parts).strip()
        except Exception as e:
            logger.error(f"Ошибка при извлечении текста из CSV {filename}: {e}")

    def extract_text_from_attachment(self, file: bytes, filename: str) -> str:
        """Извлекает текст из вложения по типу файла"""
        if not file:
            return f"Пустой файл: {filename}"

        ext = os.path.splitext(filename)[1].lower()

        try:
            if ext == '.pdf':
                return self.extract_text_from_pdf(file, filename)
            elif ext == '.docx':
                return self.extract_text_from_docx(file, filename)
            elif ext in ['.xlsx', '.xls']:
                return self.extract_text_from_excel(file, filename)
            elif ext == '.csv':
                return self.extract_text_from_csv(file, filename)
            elif ext in ['.txt', '.text', '.log']:
                try:
                    return file.decode('utf-8').strip()
                except:
                    return file.decode('cp1251', errors='ignore').strip()
            elif ext in ['.eml', '.msg']:
                # Рекурсивно парсим вложенные письма
                nested_content = self.get_email_content(file, filename)
                return f"Вложенное письмо: {filename}]\n{nested_content}"
            else:
                # Для неподдерживаемых форматов возвращаем информацию о файле
                file_size = len(file)
                return f"Бинарный файл: {filename}, размер: {file_size} байт, тип: {ext}"
        except Exception as e:
            logger.error(f"Ошибка при обработке вложения файла {filename}: {e}")
            return "Не удалось извлечь данные из вложения"

    def get_email_content(self, file: bytes, filename: str, include_attachments: bool = True) -> Tuple[str, List[Dict]]:
        """
        Главная функция: извлекает полное текстовое содержимое письма

        Args:
            file: Байты  файла
            filename: Имя файла (для определения типа)
            include_attachments: Включать ли текст из вложений

        Returns:
            Tuple[текст письма, список информации о вложениях]
        """
        attachment_info_list = []

        try:
        # Определяем тип файла по расширению
            ext = os.path.splitext(filename)[1].lower()

            if ext == '.eml':
                email_data = self.parse_eml(file, filename)
            elif ext == '.msg':
                email_data = self.parse_msg(file, filename)
            else:
                raise ValueError(f"Неподдерживаемый формат файла: {ext}. Допустимы типы файлов .eml или .msg")

            text_parts = []
            # Добавляем тело письма
            if email_data['body_plain']:
                text_parts.append(email_data['body_plain'])

            # Обрабатываем вложения
            if include_attachments and email_data['attachments']:

                for i, attachment in enumerate(email_data['attachments']):
                    if attachment.get('data'):
                        # Извлекаем текст из вложения
                        filename = attachment['filename'] or f"attachment_{i}"
                        attachment_text = self.extract_text_from_attachment(
                            attachment['data'], 
                            filename
                        )
                        # Сохраняем информацию о вложении
                        attachment_info = {
                            'filename': filename,
                            'data': attachment_text, 
                            'content_type': attachment.get('content_type', ''),
                            'size': len(attachment['data'])
                        }
                        attachment_info_list.append(attachment_info)

            # Объединяем все части в один текст
            full_text = "\n".join(text_parts)

            # Очищаем от лишних пробелов и пустых строк
            clean_text = self.clean_text(full_text)
            clean_text, urls = self.extract_and_remove_urls(clean_text)
            logger.info(f"Из электронного письма извлечено {len(clean_text)} символов")
            return {
                'subject': f"Тема письма {email_data['subject']}" if email_data['subject'] else 'Без темы',
                'from': f"От: {email_data['from']}" if email_data['from'] else 'Неизветный отправитель',
                'to': f"Кому: {email_data['to']}" if email_data['to'] else 'Неизвестный получатель',
                'date': f"Дата: {email_data['date']}" if email_data['date'] else 'Нет даты',
                "body": clean_text,
                "attachments": attachment_info_list,  # список словарей или пустой список
                "urls": urls
            }
        except Exception as e:
            logger.error(f"Ошибка при получении содержимого электронного письма из байтов: {e}")
            raise ValueError(f"Ошибка при обработке письма {filename}: {e}")


# Функции для быстрого доступа к EmailParser
def parse_email(file: bytes, filename: str, include_attachments: bool = True) -> Tuple[str, List[Dict]]:
    """
    Функция для парсинга писем вместа с вложениями

    Args:
        file: Байты файла
        filename: Имя файла (для определения типа)
        include_attachments: Включать ли текст из вложений

    Returns:
        Кортеж (текст письма, список информации о вложениях)
    """
    parser = EmailParser()
    return parser.get_email_content(file, filename, include_attachments)


def get_email_text_only(file: bytes, filename: str) -> str:
    """
    Функция для иизвлечения только текста из письма (без обработки вложений)

    Args:
        file: Байты файла
        filename: Имя файла (для определения типа)

    Returns:
        Текст письма (без текста из вложений)
    """
    parser = EmailParser()
    
    try:
        ext = os.path.splitext(filename)[1].lower()

        if ext == '.eml':
            email_data = parser.parse_eml(file, filename)
        elif ext == '.msg':
            email_data = parser.parse_msg(file, filename)
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {ext}")

        # Формируем базовый текст без вложений
        text_parts = []

        if email_data['subject']:
            text_parts.append(f"Тема: {email_data['subject']}")
        if email_data['from']:
            text_parts.append(f"От: {email_data['from']}")
        if email_data['to']:
            text_parts.append(f"Кому: {email_data['to']}")

        text_parts.append("")

        if email_data['body_plain']:
            text_parts.append(email_data['body_plain'])

        full_text = "\n".join(text_parts)
        return full_text.strip()

    except Exception as e:
        logger.error(f"Ошибка при получении текста электронного письма из байтов: {e}")
        raise ValueError(f"Ошибка при обработке письма: {e}")

def prepare_for_classification(parsed_data: dict[str, Any]) -> str:
    """
    Функция подготовки текста для модели на основе распарсенных данных
    """
    parts = []
    
    # Тема письма
    subject = parsed_data.get("subject", "")
    if subject and subject not in ["Без темы", "Ошибка парсинга"]:
        parts.append(subject)
    
    # Тело письма
    body = parsed_data.get("body", "")
    parts.append(f"Текст письма: {body}")

    # Формируем данные о вложениях
    attachments = parsed_data.get("attachments", [])
    if attachments:
        extensions = set()
        for att in attachments:
            if 'image' not in att['content_type']:
                extensions.add(att['data'][:200])
        if extensions:
            parts.append(f"Вложения: {', '.join(sorted(extensions))}")
    result = "\n".join(parts)
    
    # Если ничего не было собрано из данных, возвращаем ошибку
    if not result:
        raise ValueError('Нет данных для оценки письма')
    
    return result

